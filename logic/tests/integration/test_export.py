"""
Integration tests for the Excel export endpoint.
"""
import pytest


class TestExportToExcel:
    def test_returns_200_with_xlsx_content_type(self, client, make_application):
        make_application(company="Stripe", position="SWE")
        make_application(company="Anthropic", position="ML Eng")

        r = client.get("/export/excel")

        assert r.status_code == 200
        content_type = r.headers.get("content-type", "")
        assert "spreadsheetml" in content_type or "xlsx" in content_type

    def test_returns_xlsx_with_correct_filename_header(self, client, make_application):
        make_application()
        r = client.get("/export/excel")

        assert r.status_code == 200
        disposition = r.headers.get("content-disposition", "")
        assert "applications.xlsx" in disposition

    def test_returns_non_empty_file_content(self, client, make_application):
        make_application()
        r = client.get("/export/excel")

        assert len(r.content) > 0

    def test_export_with_no_applications_returns_200(self, client):
        """Export must still succeed even when the user has no applications."""
        r = client.get("/export/excel")
        assert r.status_code == 200

    def test_exported_file_is_valid_xlsx(self, client, make_application):
        """The bytes returned must be a valid Excel workbook."""
        import io
        from openpyxl import load_workbook

        make_application(company="Test Corp", position="Engineer")
        r = client.get("/export/excel")

        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        assert ws.title == "Applications"

        # First row should be headers
        headers = [ws.cell(1, col).value for col in range(1, len(ws[1]) + 1)]
        assert "Company" in headers
        assert "Position" in headers
        assert "Status" in headers

    def test_exported_file_contains_user_data(self, client, make_application):
        import io
        from openpyxl import load_workbook

        make_application(company="Unique Corp XYZ", position="Unique Role ABC")
        r = client.get("/export/excel")

        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active

        # Collect all cell values
        all_values = [ws.cell(row, col).value for row in ws.iter_rows() for col in range(1, ws.max_column + 1)]
        flat = [str(v) for v in all_values if v is not None]

        assert any("Unique Corp XYZ" in v for v in flat)
        assert any("Unique Role ABC" in v for v in flat)

    def test_export_requires_auth(self):
        from fastapi.testclient import TestClient
        from app.main import app

        with TestClient(app) as c:
            r = c.get("/export/excel")
        assert r.status_code == 403
